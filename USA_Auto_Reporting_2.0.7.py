import os
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton,QFileDialog, QGridLayout, QStackedWidget, QMessageBox
from PyQt5.QtGui import QFont, QIcon
import sys
import codecs


class Main_Page(QWidget):
    judge_file_open = 0

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):

        btn1 = QPushButton('대상 폴더', self)

        btn1.setMaximumWidth(180)
        btn1.setMaximumHeight(60)

        btn1.setFont(QFont('Arial', 13, QFont.Bold))
        btn1.setStyleSheet("color : white;"
                           "background-color : rgb(131, 56, 236);"
                           "border-radius : 5px;"
                           )


        btn1.clicked.connect(self.getData_raw)

        btn2 = QPushButton("Run!!", self)

        btn2.setMaximumWidth(180)
        btn2.setMaximumHeight(60)

        btn2.setFont(QFont('Arial', 13, QFont.Bold))
        btn2.setStyleSheet("color : white;"
                           "background-color : rgb(131, 56, 236);"
                           "border-radius : 5px;"
                           )

        btn2.clicked.connect(self.clickMethod)
        btn2.clicked.connect(self.run)
        grid = QGridLayout()
        grid.addWidget(btn1, 1, 1)
        grid.addWidget(btn2, 1, 3)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)
        grid.setColumnStretch(2, 1)
        grid.setColumnStretch(3, 1)
        grid.setColumnStretch(4, 1)

        self.setLayout(grid)

        self.show()

    def getData_raw(self):
        global Path_Raw
        self.judge_file_open = 1
        Path_Raw = str(QFileDialog.getExistingDirectory(self, "Select Directory"))

    def run(self):
        if self.judge_file_open == 0:
            return 0
        Path_Raw_Text = "C:\\Program Files (x86)\\ETS-Lindgren\\EMQuest\\Final Data"  #"C:\\Users\\jiyoon_kim\\Desktop\\USA_SGS_Auto\\text_data"
        num_of_check = 1000

        # 경로 지정

        raw_name_FS = os.listdir(Path_Raw + "\\" + "FS")[0:num_of_check]
        raw_name_HL = os.listdir(Path_Raw + "\\" + "HL")[0:num_of_check]
        raw_name_BHHR = os.listdir(Path_Raw + "\\" + "BHHR")[0:num_of_check]

        raw_names_FS = []
        raw_names_HL = []
        raw_names_BHHR = []


        for i in raw_name_FS:
            raw_names_FS.append(i.replace(".raw", ".txt"))

        for i in raw_name_HL:
            raw_names_HL.append(i.replace(".raw", ".txt"))

        for i in raw_name_BHHR:
            raw_names_BHHR.append(i.replace(".raw", ".txt"))

        raw_names_TOTAL = [raw_names_FS, raw_names_HL, raw_names_BHHR]

        texting = list(reversed(os.listdir(Path_Raw_Text)))
        text_names = texting[0:num_of_check]

        Band_TRP = ["66A-n5A Desense", "66A-n2A Desense", "2A-n77A Desense", "5A-n77A Desense", "66A-n77A Desense", "13A-66A-n2A", "2A-66A-n5A", "2A-13A-66A-n77A", "2A-n5A", "13A-n2A", "13A-n5A", "13A-n66A","66A-n5A", "48A-n5A", "66A-n2A", "2A-n77A", "13A-n77A", "66A-n77A"]
        Band_TIS = ["66A-n5A Desense", "66A-n2A Desense", "2A-n77A Desense", "5A-n77A Desense", "66A-n77A Desense", "13A-66A-n2A std", "13A-66A-n2A PCC", "13A-66A-n2A SCC", "2A-66A-n5A std", "2A-66A-n5A PCC", "2A-66A-n5A SCC", "2A-13A-66A-n77A std", "2A-13A-66A-n77A PCC", "2A-13A-66A-n77A SCC1", "2A-13A-66A-n77A SCC2", "2A-n5A std", "13A-n2A std", "13A-n5A std", "13A-n66A std", "48A-n5A std", "66A-n2A std", "66A-n5A std", "2A-n77A std", "13A-n77A std", "66A-n77A std"]

        middle_name_TIS = {}
        middle_name_TRP = {}

        TIS_Band_2_5 = {}
        TIS_Band_13_2 = {}
        TIS_Band_13_5 = {}
        TIS_Band_13_66 = {}
        TIS_Band_48_5 = {}
        TIS_Band_66_2 = {}
        TIS_Band_66_5 = {}
        TIS_Band_2_77 = {}
        TIS_Band_13_77 = {}
        TIS_Band_66_77 = {}
        TIS_Band_13_66_2 = {}
        TIS_Band_2_66_5 = {}
        TIS_Band_2_13_66_77 = {}
        TIS_Band_PCC_13_66_2 = {}
        TIS_Band_SCC_13_66_2 = {}
        TIS_Band_PCC_2_66_5 = {}
        TIS_Band_SCC_2_66_5 = {}
        TIS_Band_PCC_2_13_66_77 = {}
        TIS_Band_SCC1_2_13_66_77 = {}
        TIS_Band_SCC2_2_13_66_77 = {}

        TIS_Band_66_5_Desense = {}
        TIS_Band_66_2_Desense = {}
        TIS_Band_2_77_Desense = {}
        TIS_Band_5_77_Desense = {}
        TIS_Band_66_77_Desense = {}

        TRP_Band_2_5 = {}
        TRP_Band_13_2 = {}
        TRP_Band_13_5 = {}
        TRP_Band_13_66 = {}
        TRP_Band_48_5 = {}
        TRP_Band_66_2 = {}
        TRP_Band_66_5 = {}
        TRP_Band_2_77 = {}
        TRP_Band_13_77 = {}
        TRP_Band_66_77 = {}
        TRP_Band_13_66_2 = {}
        TRP_Band_2_66_5 = {}
        TRP_Band_2_13_66_77 = {}

        TRP_Band_66_5_Desense = {}
        TRP_Band_66_2_Desense = {}
        TRP_Band_2_77_Desense = {}
        TRP_Band_5_77_Desense = {}
        TRP_Band_66_77_Desense = {}

        TRP_keys = []
        TIS_keys = []

        # 텍스트에 있는 이름 저장
        for text in text_names:
            for i, config in enumerate(raw_names_TOTAL):
                for raw in config:
                    if text.replace(" .txt","") in raw:
                        with codecs.open(Path_Raw_Text + "\\" + text, 'r', encoding='utf-8', errors='ignore') as f:
                            if i == 0:
                                text_rep = text.replace(".txt", "_FS")
                            if i == 1:
                                text_rep = text.replace(".txt", "_HL")
                            if i == 2:
                                text_rep = text.replace(".txt", "_BHHR")

                            lines = f.readlines()

        # ------------------TIS 처리-----------------------------------------------------------------------------------

                            if lines[10].strip().find('TIS') >= 0:
                                middle_name_TIS[text_rep] = lines[10].strip()             # 파일명


                                TIS_Finder = "".join(s for s in lines if 'TIS (dBm)'.lower() in s.strip().lower())
                                NHPIS_45_Finder = "".join(s for s in lines if '/4 (dBm)'.lower() in s.strip().lower())
                                NHPIS_30_Finder = "".join(s for s in lines if '/6 (dBm)'.lower() in s.strip().lower())

                                if '66A-n5A' in middle_name_TIS[text_rep] and 'Desense' in middle_name_TIS[text_rep]:
                                    TIS_Band_66_5_Desense[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                                       NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2) + 1:NHPIS_45_Finder.find('\t', 24)],
                                                                       NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2) + 1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '66A-n2A' in middle_name_TIS[text_rep] and 'Desense' in middle_name_TIS[text_rep]:
                                    TIS_Band_66_2_Desense[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '2A-n77A' in middle_name_TIS[text_rep] and 'Desense' in middle_name_TIS[text_rep]:
                                    TIS_Band_2_77_Desense[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '5A-n77A' in middle_name_TIS[text_rep] and 'Desense' in middle_name_TIS[text_rep]:
                                    TIS_Band_5_77_Desense[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2) + 1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2) + 1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '66A-n77A' in middle_name_TIS[text_rep] and 'Desense' in middle_name_TIS[text_rep]:
                                    TIS_Band_66_77_Desense[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2) + 1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2) + 1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '13A-66A-n2A' in middle_name_TIS[text_rep] and 'PCC' in middle_name_TIS[text_rep]:
                                    TIS_Band_PCC_13_66_2[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '13A-66A-n2A' in middle_name_TIS[text_rep] and 'SCC' in middle_name_TIS[text_rep]:
                                    TIS_Band_SCC_13_66_2[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '13A-66A-n2A' in middle_name_TIS[text_rep]:
                                    TIS_Band_13_66_2[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '2A-66A-n5A' in middle_name_TIS[text_rep] and 'PCC' in middle_name_TIS[text_rep]:
                                    TIS_Band_PCC_2_66_5[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '2A-66A-n5A' in middle_name_TIS[text_rep] and 'SCC' in middle_name_TIS[text_rep]:
                                    TIS_Band_SCC_2_66_5[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '2A-66A-n5A' in middle_name_TIS[text_rep]:
                                    TIS_Band_2_66_5[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '2A-13A-66A-n77A' in middle_name_TIS[text_rep] and 'PCC' in middle_name_TIS[text_rep]:
                                    TIS_Band_PCC_2_13_66_77[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '2A-13A-66A-n77A' in middle_name_TIS[text_rep] and 'SCC1' in middle_name_TIS[text_rep]:
                                    TIS_Band_SCC1_2_13_66_77[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '2A-13A-66A-n77A' in middle_name_TIS[text_rep] and 'SCC2' in middle_name_TIS[text_rep]:
                                    TIS_Band_SCC2_2_13_66_77[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '2A-13A-66A-n77A' in middle_name_TIS[text_rep]:
                                    TIS_Band_2_13_66_77[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '2A-n5A' in middle_name_TIS[text_rep]:
                                    TIS_Band_2_5[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '13A-n2A' in middle_name_TIS[text_rep]:
                                    TIS_Band_13_2[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '13A-n5A' in middle_name_TIS[text_rep]:
                                    TIS_Band_13_5[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '13A-n66A' in middle_name_TIS[text_rep]:
                                    TIS_Band_13_66[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '48A-n5A' in middle_name_TIS[text_rep]:
                                    TIS_Band_48_5[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '2A-n77A' in middle_name_TIS[text_rep]:
                                    TIS_Band_2_77[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '13A-n77A' in middle_name_TIS[text_rep]:
                                    TIS_Band_13_77[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '66A-n2A' in middle_name_TIS[text_rep]:
                                    TIS_Band_66_2[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '66A-n5A' in middle_name_TIS[text_rep]:
                                    TIS_Band_66_5[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]

                                elif '66A-n77A' in middle_name_TIS[text_rep]:
                                    TIS_Band_66_77[text_rep] = [TIS_Finder[TIS_Finder.find('-'):TIS_Finder.find('\t', 14)],
                                                          NHPIS_45_Finder[NHPIS_45_Finder.find('\t', 2)+1:NHPIS_45_Finder.find('\t', 24)],
                                                          NHPIS_30_Finder[NHPIS_30_Finder.find('\t', 2)+1:NHPIS_30_Finder.find('\t', 24)]]


                                TIS_keys.append(text_rep)


        # ------------------TRP 처리-----------------------------------------------------------------------------------

                            elif lines[10].strip().find('TRP') >= 0:
                                middle_name_TRP[text_rep] = lines[10].strip()               # 파일명

                                TRP_Finder = "".join(s for s in lines if 'Tot.'.lower() in s.strip().lower())
                                NHPRP_45_Finder = "".join(s for s in lines if '/4 (dBm)'.lower() in s.strip().lower())
                                NHPRP_30_Finder = "".join(s for s in lines if '/6 (dBm)'.lower() in s.strip().lower())

                                if '66A-n5A' in middle_name_TRP[text_rep] and 'Desense' in middle_name_TRP[text_rep]:
                                    TRP_Band_66_5_Desense[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2) + 1:TRP_Finder.find('\t', 24)],
                                                                      NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2) + 1:NHPRP_45_Finder.find('\t', 24)],
                                                                      NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2) + 1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '66A-n2A' in middle_name_TRP[text_rep] and 'Desense' in middle_name_TRP[text_rep]:
                                    TRP_Band_66_2_Desense[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2) + 1:TRP_Finder.find('\t', 24)],
                                                                           NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2) + 1:NHPRP_45_Finder.find('\t', 24)],
                                                                           NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2) + 1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '2A-n77A' in middle_name_TRP[text_rep] and 'Desense' in middle_name_TRP[text_rep]:
                                    TRP_Band_2_77_Desense[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2) + 1:TRP_Finder.find('\t', 24)],
                                                                           NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2) + 1:NHPRP_45_Finder.find('\t', 24)],
                                                                           NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2) + 1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '5A-n77A' in middle_name_TRP[text_rep] and 'Desense' in middle_name_TRP[text_rep]:
                                    TRP_Band_5_77_Desense[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2) + 1:TRP_Finder.find('\t', 24)],
                                                                           NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2) + 1:NHPRP_45_Finder.find('\t', 24)],
                                                                           NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2) + 1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '66A-n77A' in middle_name_TRP[text_rep] and 'Desense' in middle_name_TRP[text_rep]:
                                    TRP_Band_66_77_Desense[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2) + 1:TRP_Finder.find('\t', 24)],
                                                                           NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2) + 1:NHPRP_45_Finder.find('\t', 24)],
                                                                           NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2) + 1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '13A-66A-n2A' in middle_name_TRP[text_rep]:
                                    TRP_Band_13_66_2[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2) + 1:TRP_Finder.find('\t', 24)],
                                                                      NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2) + 1:NHPRP_45_Finder.find('\t', 24)],
                                                                      NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2) + 1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '2A-66A-n5A' in middle_name_TRP[text_rep]:
                                    TRP_Band_2_66_5[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2)+1:TRP_Finder.find('\t', 24)],
                                                           NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2)+1:NHPRP_45_Finder.find('\t', 24)],
                                                           NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2)+1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '2A-13A-66A-n77A' in middle_name_TRP[text_rep]:
                                    TRP_Band_2_13_66_77[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2)+1:TRP_Finder.find('\t', 24)],
                                                           NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2)+1:NHPRP_45_Finder.find('\t', 24)],
                                                           NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2)+1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '2A-n5A' in middle_name_TRP[text_rep]:
                                    TRP_Band_2_5[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2)+1:TRP_Finder.find('\t', 24)],
                                                NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2)+1:NHPRP_45_Finder.find('\t', 24)],
                                                NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2)+1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '13A-n2A' in middle_name_TRP[text_rep]:
                                    TRP_Band_13_2[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2)+1:TRP_Finder.find('\t', 24)],
                                                           NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2)+1:NHPRP_45_Finder.find('\t', 24)],
                                                           NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2)+1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '13A-n5A' in middle_name_TRP[text_rep]:
                                    TRP_Band_13_5[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2)+1:TRP_Finder.find('\t', 24)],
                                                          NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2)+1:NHPRP_45_Finder.find('\t', 24)],
                                                          NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2)+1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '13A-n66A' in middle_name_TRP[text_rep]:
                                    TRP_Band_13_66[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2)+1:TRP_Finder.find('\t', 24)],
                                                          NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2)+1:NHPRP_45_Finder.find('\t', 24)],
                                                          NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2)+1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '48A-n5A' in middle_name_TRP[text_rep]:
                                    TRP_Band_48_5[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2)+1:TRP_Finder.find('\t', 24)],
                                                          NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2)+1:NHPRP_45_Finder.find('\t', 24)],
                                                          NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2)+1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '2A-n77A' in middle_name_TRP[text_rep]:
                                    TRP_Band_2_77[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2)+1:TRP_Finder.find('\t', 24)],
                                                           NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2)+1:NHPRP_45_Finder.find('\t', 24)],
                                                           NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2)+1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '13A-n77A' in middle_name_TRP[text_rep]:
                                    TRP_Band_13_77[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2)+1:TRP_Finder.find('\t', 24)],
                                                           NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2)+1:NHPRP_45_Finder.find('\t', 24)],
                                                           NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2)+1:NHPRP_30_Finder.find('\t', 24)]]
                                elif '66A-n2A' in middle_name_TRP[text_rep]:
                                    TRP_Band_66_2[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2)+1:TRP_Finder.find('\t', 24)],
                                                          NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2)+1:NHPRP_45_Finder.find('\t', 24)],
                                                          NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2)+1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '66A-n5A' in middle_name_TRP[text_rep]:
                                    TRP_Band_66_5[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2)+1:TRP_Finder.find('\t', 24)],
                                                           NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2)+1:NHPRP_45_Finder.find('\t', 24)],
                                                           NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2)+1:NHPRP_30_Finder.find('\t', 24)]]

                                elif '66A-n77A' in middle_name_TRP[text_rep]:
                                    TRP_Band_66_77[text_rep] = [TRP_Finder[TRP_Finder.find('\t', 2)+1:TRP_Finder.find('\t', 24)],
                                                           NHPRP_45_Finder[NHPRP_45_Finder.find('\t', 2)+1:NHPRP_45_Finder.find('\t', 24)],
                                                           NHPRP_30_Finder[NHPRP_30_Finder.find('\t', 2)+1:NHPRP_30_Finder.find('\t', 24)]]

                                TRP_keys.append(text_rep)

        # -------------------------------------------------------------------------------------------

        # TRP,TIS 엑셀 작성

        Excel_TIS_Band = {"66A-n5A Desense":TIS_Band_66_5_Desense,"66A-n2A Desense":TIS_Band_66_2_Desense,"2A-n77A Desense":TIS_Band_2_77_Desense,"5A-n77A Desense":TIS_Band_5_77_Desense,"66A-n77A Desense":TIS_Band_66_77_Desense, "2A-n5A std":TIS_Band_2_5, "13A-n2A std":TIS_Band_13_2, "13A-n5A std":TIS_Band_13_5, "13A-n66A std":TIS_Band_13_66, "48A-n5A std":TIS_Band_48_5, "66A-n2A std":TIS_Band_66_2, "66A-n5A std":TIS_Band_66_5, "2A-n77A std":TIS_Band_2_77, "13A-n77A std":TIS_Band_13_77, "66A-n77A std":TIS_Band_66_77,
                          "13A-66A-n2A std":TIS_Band_13_66_2, "13A-66A-n2A PCC":TIS_Band_PCC_13_66_2, "13A-66A-n2A SCC":TIS_Band_SCC_13_66_2, "2A-66A-n5A std":TIS_Band_2_66_5, "2A-66A-n5A PCC":TIS_Band_PCC_2_66_5, "2A-66A-n5A SCC":TIS_Band_SCC_2_66_5, "2A-13A-66A-n77A std":TIS_Band_2_13_66_77, "2A-13A-66A-n77A PCC":TIS_Band_PCC_2_13_66_77, "2A-13A-66A-n77A SCC1":TIS_Band_SCC1_2_13_66_77, "2A-13A-66A-n77A SCC2":TIS_Band_SCC2_2_13_66_77}


        Excel_TRP_Band = {"66A-n5A Desense":TRP_Band_66_5_Desense,"66A-n2A Desense":TRP_Band_66_2_Desense,"2A-n77A Desense":TRP_Band_2_77_Desense,"5A-n77A Desense":TRP_Band_5_77_Desense,"66A-n77A Desense":TRP_Band_66_77_Desense, "2A-n5A":TRP_Band_2_5, "13A-n2A":TRP_Band_13_2, "13A-n5A":TRP_Band_13_5, "13A-n66A":TRP_Band_13_66, "48A-n5A":TRP_Band_48_5, "66A-n2A":TRP_Band_66_2, "66A-n5A":TRP_Band_66_5,
                    "2A-n77A":TRP_Band_2_77, "13A-n77A":TRP_Band_13_77, "66A-n77A":TRP_Band_66_77, "13A-66A-n2A":TRP_Band_13_66_2, "2A-66A-n5A":TRP_Band_2_66_5, "2A-13A-66A-n77A":TRP_Band_2_13_66_77}

        if TRP_keys:
            load_TRP_NR = load_workbook("C:\\Program Files (x86)\ETS-Lindgren\\EMQuest\\Auto Reporting_Template\\OTA TRP Result_NR.XLSX")    #"C:\\Users\\jiyoon_kim\\Desktop\\USA_SGS_Auto\\Auto Reporting_Template\\OTA TRP Result_NR.XLSX"
            load_TRP_LTE = load_workbook("C:\\Program Files (x86)\ETS-Lindgren\\EMQuest\\Auto Reporting_Template\\OTA TRP Result_LTE.XLSX")

            load_TRP_NR_Band_Sheet_2_5 = load_TRP_NR['2A-n5A_TRP']
            load_TRP_NR_Band_Sheet_13_2 = load_TRP_NR['13A-n2A_TRP']
            load_TRP_NR_Band_Sheet_13_5 = load_TRP_NR['13A-n5A_TRP']
            load_TRP_NR_Band_Sheet_13_66 = load_TRP_NR['13A-n66A_TRP']
            load_TRP_NR_Band_Sheet_48_5 = load_TRP_NR['48A-n5A_TRP']
            load_TRP_NR_Band_Sheet_66_2 = load_TRP_NR['66A-n2A_TRP']
            load_TRP_NR_Band_Sheet_66_5 = load_TRP_NR['66A-n5A_TRP']
            load_TRP_NR_Band_Sheet_2_77 = load_TRP_NR['2A-n77A_TRP']
            load_TRP_NR_Band_Sheet_13_77 = load_TRP_NR['13A-n77A_TRP']
            load_TRP_NR_Band_Sheet_66_77 = load_TRP_NR['66A-n77A_TRP']
            load_TRP_NR_Band_Sheet_13_66_2 = load_TRP_NR['13A-66A-n2A_TRP']
            load_TRP_NR_Band_Sheet_2_66_5 = load_TRP_NR['2A-66A-n5A_TRP']
            load_TRP_NR_Band_Sheet_2_13_66_77 = load_TRP_NR['2A-13A-66A-n77A_TRP']
            load_TRP_NR_Band_Sheet_66_5_Desense = load_TRP_NR['66A-n5A_TRP_Desense']
            load_TRP_NR_Band_Sheet_66_2_Desense = load_TRP_NR['66A-n2A_TRP_Desense']
            load_TRP_NR_Band_Sheet_2_77_Desense = load_TRP_NR['2A-n77A_TRP_Desense']
            load_TRP_NR_Band_Sheet_5_77_Desense = load_TRP_NR['5A-n77A_TRP_Desense']
            load_TRP_NR_Band_Sheet_66_77_Desense = load_TRP_NR['66A-n77A_TRP_Desense']


            load_TRP_LTE_Band_Sheet_2_5 = load_TRP_LTE['2A-n5A_TRP']
            load_TRP_LTE_Band_Sheet_13_2 = load_TRP_LTE['13A-n2A_TRP']
            load_TRP_LTE_Band_Sheet_13_5 = load_TRP_LTE['13A-n5A_TRP']
            load_TRP_LTE_Band_Sheet_13_66 = load_TRP_LTE['13A-n66A_TRP']
            load_TRP_LTE_Band_Sheet_48_5 = load_TRP_LTE['48A-n5A_TRP']
            load_TRP_LTE_Band_Sheet_66_2 = load_TRP_LTE['66A-n2A_TRP']
            load_TRP_LTE_Band_Sheet_66_5 = load_TRP_LTE['66A-n5A_TRP']
            load_TRP_LTE_Band_Sheet_2_77 = load_TRP_LTE['2A-n77A_TRP']
            load_TRP_LTE_Band_Sheet_13_77 = load_TRP_LTE['13A-n77A_TRP']
            load_TRP_LTE_Band_Sheet_66_77 = load_TRP_LTE['66A-n77A_TRP']
            load_TRP_LTE_Band_Sheet_13_66_2 = load_TRP_LTE['13A-66A-n2A_TRP']
            load_TRP_LTE_Band_Sheet_2_66_5 = load_TRP_LTE['2A-66A-n5A_TRP']
            load_TRP_LTE_Band_Sheet_2_13_66_77 = load_TRP_LTE['2A-13A-66A-n77A_TRP']
            load_TRP_LTE_Band_Sheet_66_5_Desense = load_TRP_LTE['66A-n5A_TRP_Desense']
            load_TRP_LTE_Band_Sheet_66_2_Desense = load_TRP_LTE['66A-n2A_TRP_Desense']
            load_TRP_LTE_Band_Sheet_2_77_Desense = load_TRP_LTE['2A-n77A_TRP_Desense']
            load_TRP_LTE_Band_Sheet_5_77_Desense = load_TRP_LTE['5A-n77A_TRP_Desense']
            load_TRP_LTE_Band_Sheet_66_77_Desense = load_TRP_LTE['66A-n77A_TRP_Desense']

            load_TRP_NR_Key = {"66A-n5A Desense" : load_TRP_NR_Band_Sheet_66_5_Desense, "66A-n2A Desense" : load_TRP_NR_Band_Sheet_66_2_Desense, "2A-n77A Desense" : load_TRP_NR_Band_Sheet_2_77_Desense, "5A-n77A Desense" : load_TRP_NR_Band_Sheet_5_77_Desense, "66A-n77A Desense" : load_TRP_NR_Band_Sheet_66_77_Desense, "2A-n5A": load_TRP_NR_Band_Sheet_2_5, "13A-n2A": load_TRP_NR_Band_Sheet_13_2, "13A-n5A": load_TRP_NR_Band_Sheet_13_5, "13A-n66A": load_TRP_NR_Band_Sheet_13_66, "48A-n5A": load_TRP_NR_Band_Sheet_48_5, "66A-n2A": load_TRP_NR_Band_Sheet_66_2,
                               "66A-n5A": load_TRP_NR_Band_Sheet_66_5, "2A-n77A": load_TRP_NR_Band_Sheet_2_77, "13A-n77A": load_TRP_NR_Band_Sheet_13_77, "66A-n77A": load_TRP_NR_Band_Sheet_66_77, "13A-66A-n2A": load_TRP_NR_Band_Sheet_13_66_2,
                               "2A-66A-n5A": load_TRP_NR_Band_Sheet_2_66_5, "2A-13A-66A-n77A": load_TRP_NR_Band_Sheet_2_13_66_77}

            load_TRP_LTE_Key = {"66A-n5A Desense" : load_TRP_LTE_Band_Sheet_66_5_Desense, "66A-n2A Desense" : load_TRP_LTE_Band_Sheet_66_2_Desense, "2A-n77A Desense" : load_TRP_LTE_Band_Sheet_2_77_Desense, "5A-n77A Desense" : load_TRP_LTE_Band_Sheet_5_77_Desense, "66A-n77A Desense" : load_TRP_LTE_Band_Sheet_66_77_Desense, "2A-n5A": load_TRP_LTE_Band_Sheet_2_5, "13A-n2A": load_TRP_LTE_Band_Sheet_13_2, "13A-n5A": load_TRP_LTE_Band_Sheet_13_5, "13A-n66A": load_TRP_LTE_Band_Sheet_13_66, "48A-n5A": load_TRP_LTE_Band_Sheet_48_5, "66A-n2A": load_TRP_LTE_Band_Sheet_66_2,
                               "66A-n5A": load_TRP_LTE_Band_Sheet_66_5, "2A-n77A": load_TRP_LTE_Band_Sheet_2_77, "13A-n77A": load_TRP_LTE_Band_Sheet_13_77, "66A-n77A": load_TRP_LTE_Band_Sheet_66_77, "13A-66A-n2A": load_TRP_LTE_Band_Sheet_13_66_2, "2A-66A-n5A": load_TRP_LTE_Band_Sheet_2_66_5, "2A-13A-66A-n77A": load_TRP_LTE_Band_Sheet_2_13_66_77}

            for TRP_key in TRP_keys:
                for i in Band_TRP:
                    if 'FS' in TRP_key and 'NR' in middle_name_TRP[TRP_key] and TRP_key in Excel_TRP_Band[i] and 'Desense' in middle_name_TRP[TRP_key] and '66A-n5A' in middle_name_TRP[TRP_key]:
                        if 'Low' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['E6'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['J6'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['O6'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'Mid' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['E7'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['J7'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['O7'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'High' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['E8'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['J8'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['O8'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                    elif 'FS' in TRP_key and 'LTE' in middle_name_TRP[TRP_key] and TRP_key in Excel_TRP_Band[i] and 'Desense' in middle_name_TRP[TRP_key] and '66A-n5A' in middle_name_TRP[TRP_key]:
                        if 'Low' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['E6'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['J6'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['O6'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'Mid' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['E7'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['J7'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['O7'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'High' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['E8'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['J8'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['O8'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                    elif 'FS' in TRP_key and 'NR' in middle_name_TRP[TRP_key] and TRP_key in Excel_TRP_Band[i] and 'Desense' in middle_name_TRP[TRP_key] and '2A-n77A' in middle_name_TRP[TRP_key]:
                        if '10MHz' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['E6'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['J6'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['O6'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif '20MHz' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['E7'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['J7'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['O7'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                    elif 'FS' in TRP_key and 'LTE' in middle_name_TRP[TRP_key] and TRP_key in Excel_TRP_Band[i] and 'Desense' in middle_name_TRP[TRP_key] and '2A-n77A' in middle_name_TRP[TRP_key]:
                        if '10MHz' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['E6'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['J6'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['O6'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif '20MHz' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['E7'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['J7'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['O7'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                    elif 'FS' in TRP_key and 'NR' in middle_name_TRP[TRP_key] and TRP_key in Excel_TRP_Band[i]:
                        if 'Low' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['E6'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['J6'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['O6'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'Mid' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['E7'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['J7'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['O7'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'High' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['E8'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['J8'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['O8'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                    elif 'FS' in TRP_key and 'LTE' in middle_name_TRP[TRP_key] and TRP_key in Excel_TRP_Band[i]:
                        if 'Low' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['E6'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['J6'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['O6'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'Mid' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['E7'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['J7'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['O7'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'High' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['E8'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['J8'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['O8'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                    elif 'HL' in TRP_key and 'NR' in middle_name_TRP[TRP_key] and TRP_key in Excel_TRP_Band[i]:
                        if 'Low' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['F6'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['K6'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['P6'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'Mid' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['F7'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['K7'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['P7'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'High' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['F8'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['K8'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['P8'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                    elif 'HL' in TRP_key and 'LTE' in middle_name_TRP[TRP_key] and TRP_key in Excel_TRP_Band[i]:
                        if 'Low' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['F6'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['K6'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['P6'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'Mid' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['F7'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['K7'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['P7'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'High' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['F8'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['K8'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['P8'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                    elif 'BHHR' in TRP_key and 'NR' in middle_name_TRP[TRP_key] and TRP_key in Excel_TRP_Band[i]:
                        if 'Low' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['I6'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['N6'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['S6'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'Mid' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['I7'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['N7'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['S7'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'High' in middle_name_TRP[TRP_key]:
                            load_TRP_NR_Key[i]['I8'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['N8'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_NR_Key[i]['S8'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                    elif 'BHHR' in TRP_key and 'LTE' in middle_name_TRP[TRP_key] and TRP_key in Excel_TRP_Band[i]:
                        if 'Low' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['I6'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['N6'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['S6'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'Mid' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['I7'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['N7'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['S7'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

                        elif 'High' in middle_name_TRP[TRP_key]:
                            load_TRP_LTE_Key[i]['I8'] = round(float(Excel_TRP_Band[i][TRP_key][0]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['N8'] = round(float(Excel_TRP_Band[i][TRP_key][1]) + 0.000001, 2)
                            load_TRP_LTE_Key[i]['S8'] = round(float(Excel_TRP_Band[i][TRP_key][2]) + 0.000001, 2)

            load_TRP_NR.save(Path_Raw +"\\" + "Result_NR_TRP.XLSX")
            load_TRP_LTE.save(Path_Raw + "\\" + "Result_LTE_TRP.XLSX")

        if TIS_keys:
           load_TIS_NR = load_workbook("C:\\Program Files (x86)\ETS-Lindgren\\EMQuest\\Auto Reporting_Template\\OTA TIS Result_NR.XLSX")
           load_TIS_LTE = load_workbook("C:\\Program Files (x86)\ETS-Lindgren\\EMQuest\\Auto Reporting_Template\\OTA TIS Result_LTE.XLSX")

           load_TIS_NR_Band_Sheet_2_5 = load_TIS_NR['2A-n5A_TIS']
           load_TIS_NR_Band_Sheet_13_2 = load_TIS_NR['13A-n2A_TIS']
           load_TIS_NR_Band_Sheet_13_5 = load_TIS_NR['13A-n5A_TIS']
           load_TIS_NR_Band_Sheet_13_66 = load_TIS_NR['13A-n66A_TIS']
           load_TIS_NR_Band_Sheet_48_5 = load_TIS_NR['48A-n5A_TIS']
           load_TIS_NR_Band_Sheet_66_2 = load_TIS_NR['66A-n2A_TIS']
           load_TIS_NR_Band_Sheet_66_5 = load_TIS_NR['66A-n5A_TIS']
           load_TIS_NR_Band_Sheet_2_77 = load_TIS_NR['2A-n77A_TIS']
           load_TIS_NR_Band_Sheet_13_77 = load_TIS_NR['13A-n77A_TIS']
           load_TIS_NR_Band_Sheet_66_77 = load_TIS_NR['66A-n77A_TIS']
           load_TIS_NR_Band_Sheet_13_66_2 = load_TIS_NR['13A-66A-n2A_TIS']
           load_TIS_NR_Band_Sheet_2_66_5 = load_TIS_NR['2A-66A-n5A_TIS']
           load_TIS_NR_Band_Sheet_2_13_66_77 = load_TIS_NR['2A-13A-66A-n77A_TIS']
           load_TIS_NR_Band_Sheet_66_5_Desense = load_TIS_NR['66A-n5A_TIS_Desense']
           load_TIS_NR_Band_Sheet_66_2_Desense = load_TIS_NR['66A-n2A_TIS_Desense']
           load_TIS_NR_Band_Sheet_2_77_Desense = load_TIS_NR['2A-n77A_TIS_Desense']
           load_TIS_NR_Band_Sheet_5_77_Desense = load_TIS_NR['5A-n77A_TIS_Desense']
           load_TIS_NR_Band_Sheet_66_77_Desense = load_TIS_NR['66A-n77A_TIS_Desense']

           load_TIS_LTE_Band_Sheet_2_5 = load_TIS_LTE['2A-n5A_TIS']
           load_TIS_LTE_Band_Sheet_13_2 = load_TIS_LTE['13A-n2A_TIS']
           load_TIS_LTE_Band_Sheet_13_5 = load_TIS_LTE['13A-n5A_TIS']
           load_TIS_LTE_Band_Sheet_13_66 = load_TIS_LTE['13A-n66A_TIS']
           load_TIS_LTE_Band_Sheet_48_5 = load_TIS_LTE['48A-n5A_TIS']
           load_TIS_LTE_Band_Sheet_66_2 = load_TIS_LTE['66A-n2A_TIS']
           load_TIS_LTE_Band_Sheet_66_5 = load_TIS_LTE['66A-n5A_TIS']
           load_TIS_LTE_Band_Sheet_2_77 = load_TIS_LTE['2A-n77A_TIS']
           load_TIS_LTE_Band_Sheet_13_77 = load_TIS_LTE['13A-n77A_TIS']
           load_TIS_LTE_Band_Sheet_66_77 = load_TIS_LTE['66A-n77A_TIS']
           load_TIS_LTE_Band_Sheet_PCC_13_66_2 = load_TIS_LTE['13A-66A-n2A_PCC_TIS']
           load_TIS_LTE_Band_Sheet_SCC_13_66_2 = load_TIS_LTE['13A-66A-n2A_SCC_TIS']
           load_TIS_LTE_Band_Sheet_PCC_2_66_5 = load_TIS_LTE['2A-66A-n5A_PCC_TIS']
           load_TIS_LTE_Band_Sheet_SCC_2_66_5 = load_TIS_LTE['2A-66A-n5A_SCC_TIS']
           load_TIS_LTE_Band_Sheet_PCC_2_13_66_77 = load_TIS_LTE['2A-13A-66A-n77A_PCC_TIS']
           load_TIS_LTE_Band_Sheet_SCC1_2_13_66_77 = load_TIS_LTE['2A-13A-66A-n77A_SCC1_TIS']
           load_TIS_LTE_Band_Sheet_SCC2_2_13_66_77 = load_TIS_LTE['2A-13A-66A-n77A_SCC2_TIS']
           load_TIS_LTE_Band_Sheet_66_5_Desense = load_TIS_LTE['66A-n5A_TIS_Desense']
           load_TIS_LTE_Band_Sheet_66_2_Desense = load_TIS_LTE['66A-n2A_TIS_Desense']
           load_TIS_LTE_Band_Sheet_2_77_Desense = load_TIS_LTE['2A-n77A_TIS_Desense']
           load_TIS_LTE_Band_Sheet_5_77_Desense = load_TIS_LTE['5A-n77A_TIS_Desense']
           load_TIS_LTE_Band_Sheet_66_77_Desense = load_TIS_LTE['66A-n77A_TIS_Desense']

           load_TIS_NR_Key = {"66A-n5A Desense" : load_TIS_NR_Band_Sheet_66_5_Desense, "66A-n2A Desense" : load_TIS_NR_Band_Sheet_66_2_Desense, "2A-n77A Desense" : load_TIS_NR_Band_Sheet_2_77_Desense, "5A-n77A Desense" : load_TIS_NR_Band_Sheet_5_77_Desense, "66A-n77A Desense" : load_TIS_NR_Band_Sheet_66_77_Desense, "2A-n5A std": load_TIS_NR_Band_Sheet_2_5, "13A-n2A std": load_TIS_NR_Band_Sheet_13_2, "13A-n5A std": load_TIS_NR_Band_Sheet_13_5, "13A-n66A std": load_TIS_NR_Band_Sheet_13_66, "48A-n5A std": load_TIS_NR_Band_Sheet_48_5, "66A-n2A std": load_TIS_NR_Band_Sheet_66_2,
                              "66A-n5A std": load_TIS_NR_Band_Sheet_66_5, "2A-n77A std": load_TIS_NR_Band_Sheet_2_77, "13A-n77A std": load_TIS_NR_Band_Sheet_13_77, "66A-n77A std": load_TIS_NR_Band_Sheet_66_77, "13A-66A-n2A std": load_TIS_NR_Band_Sheet_13_66_2, "2A-66A-n5A std": load_TIS_NR_Band_Sheet_2_66_5, "2A-13A-66A-n77A std": load_TIS_NR_Band_Sheet_2_13_66_77}

           load_TIS_LTE_Key = {"66A-n5A Desense" : load_TIS_LTE_Band_Sheet_66_5_Desense, "66A-n2A Desense" : load_TIS_LTE_Band_Sheet_66_2_Desense, "2A-n77A Desense" : load_TIS_LTE_Band_Sheet_2_77_Desense, "5A-n77A Desense" : load_TIS_LTE_Band_Sheet_5_77_Desense, "66A-n77A Desense" : load_TIS_LTE_Band_Sheet_66_77_Desense, "2A-n5A std": load_TIS_LTE_Band_Sheet_2_5, "13A-n2A std": load_TIS_LTE_Band_Sheet_13_2, "13A-n5A std": load_TIS_LTE_Band_Sheet_13_5, "13A-n66A std": load_TIS_LTE_Band_Sheet_13_66, "48A-n5A std": load_TIS_LTE_Band_Sheet_48_5, "66A-n2A std": load_TIS_LTE_Band_Sheet_66_2,
                              "66A-n5A std": load_TIS_LTE_Band_Sheet_66_5, "2A-n77A std": load_TIS_LTE_Band_Sheet_2_77, "13A-n77A std": load_TIS_LTE_Band_Sheet_13_77, "66A-n77A std": load_TIS_LTE_Band_Sheet_66_77, "13A-66A-n2A PCC": load_TIS_LTE_Band_Sheet_PCC_13_66_2, "13A-66A-n2A SCC": load_TIS_LTE_Band_Sheet_SCC_13_66_2, "2A-66A-n5A PCC": load_TIS_LTE_Band_Sheet_PCC_2_66_5,
                              "2A-66A-n5A SCC": load_TIS_LTE_Band_Sheet_SCC_2_66_5, "2A-13A-66A-n77A PCC": load_TIS_LTE_Band_Sheet_PCC_2_13_66_77, "2A-13A-66A-n77A SCC1": load_TIS_LTE_Band_Sheet_SCC1_2_13_66_77, "2A-13A-66A-n77A SCC2": load_TIS_LTE_Band_Sheet_SCC2_2_13_66_77}

           for TIS_key in TIS_keys:
               for i in Band_TIS:
                   if 'FS' in TIS_key and 'NR' in middle_name_TIS[TIS_key] and TIS_key in Excel_TIS_Band[i] and 'Desense' in middle_name_TIS[TIS_key] and '66A-n5A' in middle_name_TIS[TIS_key]:
                       if 'Low' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['F6'] = round(float(Excel_TIS_Band[i][TIS_key][0]) + 0.000001, 2)
                           load_TIS_NR_Key[i]['K6'] = round(float(Excel_TIS_Band[i][TIS_key][1]) + 0.000001, 2)
                           load_TIS_NR_Key[i]['P6'] = round(float(Excel_TIS_Band[i][TIS_key][2]) + 0.000001, 2)

                       elif 'Mid' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['F7'] = round(float(Excel_TIS_Band[i][TIS_key][0]) + 0.000001, 2)
                           load_TIS_NR_Key[i]['K7'] = round(float(Excel_TIS_Band[i][TIS_key][1]) + 0.000001, 2)
                           load_TIS_NR_Key[i]['P7'] = round(float(Excel_TIS_Band[i][TIS_key][2]) + 0.000001, 2)

                       elif 'High' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['F8'] = round(float(Excel_TIS_Band[i][TIS_key][0]) + 0.000001, 2)
                           load_TIS_NR_Key[i]['K8'] = round(float(Excel_TIS_Band[i][TIS_key][1]) + 0.000001, 2)
                           load_TIS_NR_Key[i]['P8'] = round(float(Excel_TIS_Band[i][TIS_key][2]) + 0.000001, 2)

                   elif 'FS' in TIS_key and 'LTE' in middle_name_TIS[TIS_key] and TIS_key in Excel_TIS_Band[i] and 'Desense' in middle_name_TIS[TIS_key] and '66A-n5A' in middle_name_TIS[TIS_key]:
                       if 'Low' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['F6'] = round(float(Excel_TIS_Band[i][TIS_key][0]) + 0.000001, 2)
                           load_TIS_LTE_Key[i]['K6'] = round(float(Excel_TIS_Band[i][TIS_key][1]) + 0.000001, 2)
                           load_TRP_LTE_Key[i]['P6'] = round(float(Excel_TIS_Band[i][TIS_key][2]) + 0.000001, 2)

                       elif 'Mid' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['F7'] = round(float(Excel_TIS_Band[i][TIS_key][0]) + 0.000001, 2)
                           load_TIS_LTE_Key[i]['K7'] = round(float(Excel_TIS_Band[i][TIS_key][1]) + 0.000001, 2)
                           load_TIS_LTE_Key[i]['P7'] = round(float(Excel_TIS_Band[i][TIS_key][2]) + 0.000001, 2)

                       elif 'High' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['F8'] = round(float(Excel_TIS_Band[i][TIS_key][0]) + 0.000001, 2)
                           load_TIS_LTE_Key[i]['K8'] = round(float(Excel_TIS_Band[i][TIS_key][1]) + 0.000001, 2)
                           load_TIS_LTE_Key[i]['P8'] = round(float(Excel_TIS_Band[i][TIS_key][2]) + 0.000001, 2)

                   elif 'FS' in TIS_key and 'NR' in middle_name_TIS[TIS_key] and TIS_key in Excel_TIS_Band[i] and 'Desense' in middle_name_TIS[TIS_key] and '2A-n77A' in middle_name_TIS[TIS_key]:
                       if '10MHz' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['F6'] = round(float(Excel_TIS_Band[i][TIS_key][0]) + 0.000001, 2)
                           load_TIS_NR_Key[i]['K6'] = round(float(Excel_TIS_Band[i][TIS_key][1]) + 0.000001, 2)
                           load_TIS_NR_Key[i]['P6'] = round(float(Excel_TIS_Band[i][TIS_key][2]) + 0.000001, 2)

                       elif '20MHz' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['F7'] = round(float(Excel_TIS_Band[i][TIS_key][0]) + 0.000001, 2)
                           load_TIS_NR_Key[i]['K7'] = round(float(Excel_TIS_Band[i][TIS_key][1]) + 0.000001, 2)
                           load_TIS_NR_Key[i]['P7'] = round(float(Excel_TIS_Band[i][TIS_key][2]) + 0.000001, 2)

                   elif 'FS' in TIS_key and 'LTE' in middle_name_TIS[TIS_key] and TIS_key in Excel_TIS_Band[i] and 'Desense' in middle_name_TIS[TIS_key] and '2A-n77A' in middle_name_TIS[TIS_key]:
                       if '10MHz' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['F6'] = round(float(Excel_TIS_Band[i][TIS_key][0]) + 0.000001, 2)
                           load_TIS_LTE_Key[i]['K6'] = round(float(Excel_TIS_Band[i][TIS_key][1]) + 0.000001, 2)
                           load_TIS_LTE_Key[i]['P6'] = round(float(Excel_TIS_Band[i][TIS_key][2]) + 0.000001, 2)

                       elif '20MHz' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['F7'] = round(float(Excel_TIS_Band[i][TIS_key][0]) + 0.000001, 2)
                           load_TIS_LTE_Key[i]['K7'] = round(float(Excel_TIS_Band[i][TIS_key][1]) + 0.000001, 2)
                           load_TIS_LTE_Key[i]['P7'] = round(float(Excel_TIS_Band[i][TIS_key][2]) + 0.000001, 2)

                   elif 'FS' in TIS_key and 'NR' in middle_name_TIS[TIS_key] and TIS_key in Excel_TIS_Band[i]:
                       if 'Low' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['F6'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['K6'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['P6'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                       elif 'Mid' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['F7'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['K7'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['P7'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                       elif 'High' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['F8'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['K8'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['P8'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                   elif 'FS' in TIS_key and 'LTE' in middle_name_TIS[TIS_key] and TIS_key in Excel_TIS_Band[i]:
                       if 'Low' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['F6'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['K6'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['P6'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                       elif 'Mid' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['F7'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['K7'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['P7'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                       elif 'High' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['F8'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['K8'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['P8'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                   elif 'BHHR' in TIS_key and 'NR' in middle_name_TIS[TIS_key] and TIS_key in Excel_TIS_Band[i]:
                       if 'Low' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['J6'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['O6'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['T6'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                       elif 'Mid' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['J7'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['O7'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['T7'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                       elif 'High' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['J8'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['O8'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['T8'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                   elif 'BHHR' in TIS_key and 'LTE' in middle_name_TIS[TIS_key] and TIS_key in Excel_TIS_Band[i]:
                       if 'Low' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['J6'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['O6'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['T6'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                       elif 'Mid' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['J7'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['O7'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['T7'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                       elif 'High' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['J8'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['O8'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['T8'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                   elif 'HL' in TIS_key and 'NR' in middle_name_TIS[TIS_key] and TIS_key in Excel_TIS_Band[i]:
                       if 'Low' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['G6'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['L6'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['Q6'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                       elif 'Mid' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['G7'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['L7'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['Q7'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                       elif 'High' in middle_name_TIS[TIS_key]:
                           load_TIS_NR_Key[i]['G8'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['L8'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_NR_Key[i]['Q8'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                   elif 'HL' in TIS_key and 'LTE' in middle_name_TIS[TIS_key] and TIS_key in Excel_TIS_Band[i]:
                       if 'Low' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['G6'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['L6'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['Q6'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                       elif 'Mid' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['G7'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['L7'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['Q7'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

                       elif 'High' in middle_name_TIS[TIS_key]:
                           load_TIS_LTE_Key[i]['G8'] = round(float(Excel_TIS_Band[i][TIS_key][0]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['L8'] = round(float(Excel_TIS_Band[i][TIS_key][1]) - 0.000001, 2)
                           load_TIS_LTE_Key[i]['Q8'] = round(float(Excel_TIS_Band[i][TIS_key][2]) - 0.000001, 2)

           load_TIS_NR.save(Path_Raw +"\\"+ "Result_NR_TIS.XLSX")
           load_TIS_LTE.save(Path_Raw +"\\"+ "Result_LTE_TIS.XLSX")


        #-------------------------------------------------------------------------------------------------------------

    def clickMethod(self):                                                                     # 폴더 지정 안했을 때 에러 처리
        if self.judge_file_open == 0:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("에러")
            msg.setInformativeText("폴더를 선택하세요.")
            msg.setWindowTitle("Error")
            msg.exec_()

        else:
            QMessageBox.information(self,"Complete","완료되었습니다.")

class Vari_QStackedWidget(QStackedWidget):

    def closeEvent(self, event):
        reply = QMessageBox.question(self, 'Message', '종료하시겠습니까?',
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)

        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()




if __name__ == "__main__":


    app = QApplication(sys.argv)
    ex = Main_Page()
    widget = Vari_QStackedWidget()
    widget.addWidget(ex)

    widget.setWindowTitle("SGS Reporting Auto_jiyoonkim")
    widget.setWindowIcon(QIcon("LOGO.ico"))
    widget.resize(1000, 500)

    widget.show()
    sys.exit(app.exec_())
